#!/usr/bin/env python3
"""
Generate DCF (Discounted Cash Flow) Valuation Excel Model

Usage:
    python3 generate-dcf.py --company "Company Name" --years 5 --output dcf.xlsx
    python3 generate-dcf.py --company "Company Name" --wacc 9.5 --growth 2.5 --output dcf.xlsx
"""

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)


# Financial color scheme
COLORS = {
    'input_blue': '0000FF',
    'calc_black': '000000',
    'link_green': '008000',
    'header_fill': 'D9E1F2',
    'subheader_fill': 'E7E6E6',
    'alternate_fill': 'F2F2F2',
    'highlight_fill': 'FFF2CC',
}


def create_styles(wb: Workbook):
    """Create named styles for the workbook."""
    
    # Input style (blue text)
    input_style = NamedStyle(name='input_style')
    input_style.font = Font(color=COLORS['input_blue'], size=10)
    input_style.number_format = '#,##0.0'
    wb.add_named_style(input_style)
    
    # Calc style (black text)
    calc_style = NamedStyle(name='calc_style')
    calc_style.font = Font(color=COLORS['calc_black'], size=10)
    calc_style.number_format = '#,##0.0'
    wb.add_named_style(calc_style)
    
    # Percent style
    percent_style = NamedStyle(name='percent_style')
    percent_style.font = Font(color=COLORS['input_blue'], size=10)
    percent_style.number_format = '0.0%'
    wb.add_named_style(percent_style)
    
    # Currency style (millions)
    currency_style = NamedStyle(name='currency_style')
    currency_style.font = Font(color=COLORS['calc_black'], size=10)
    currency_style.number_format = '$#,##0.0,,"M"'
    wb.add_named_style(currency_style)
    
    # Currency input style
    currency_input_style = NamedStyle(name='currency_input_style')
    currency_input_style.font = Font(color=COLORS['input_blue'], size=10)
    currency_input_style.number_format = '$#,##0.0,,"M"'
    wb.add_named_style(currency_input_style)
    
    # Header style
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(bold=True, color='FFFFFF', size=11)
    header_style.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wb.add_named_style(header_style)
    
    # Subheader style
    subheader_style = NamedStyle(name='subheader_style')
    subheader_style.font = Font(bold=True, size=10)
    subheader_style.fill = PatternFill(start_color=COLORS['subheader_fill'], 
                                       end_color=COLORS['subheader_fill'], fill_type='solid')
    wb.add_named_style(subheader_style)
    
    # Highlight style (for key outputs)
    highlight_style = NamedStyle(name='highlight_style')
    highlight_style.font = Font(bold=True, size=11)
    highlight_style.fill = PatternFill(start_color=COLORS['highlight_fill'],
                                       end_color=COLORS['highlight_fill'], fill_type='solid')
    highlight_style.number_format = '$#,##0.0,,"M"'
    wb.add_named_style(highlight_style)
    
    return input_style, calc_style


def setup_sheet(ws, title: str):
    """Setup worksheet with title and basic formatting."""
    ws.title = title
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    for col in range(3, 15):
        ws.column_dimensions[get_column_letter(col)].width = 14
    
    # Title
    ws['A1'] = 'DISCOUNTED CASH FLOW (DCF) VALUATION'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:M1')


def add_assumptions_section(ws, start_row: int, wacc: float, growth: float, years: int):
    """Add the key assumptions section."""
    row = start_row
    
    # Section header
    ws[f'A{row}'] = 'KEY ASSUMPTIONS'
    ws[f'A{row}'].style = 'subheader_style'
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # WACC section
    ws[f'A{row}'] = 'Weighted Average Cost of Capital (WACC)'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    wacc_items = [
        ('Risk-free Rate', 0.04, 'percent_style'),
        ('Market Risk Premium', 0.055, 'percent_style'),
        ('Beta', 1.0, 'input_style'),
        ('Cost of Equity', '', 'percent_style'),
        ('Cost of Debt', 0.05, 'percent_style'),
        ('Tax Rate', 0.21, 'percent_style'),
        ('Target Debt/Capital', 0.30, 'percent_style'),
        ('WACC', wacc/100, 'percent_style'),
    ]
    
    for item, value, style in wacc_items:
        ws[f'A{row}'] = f'  {item}'
        cell = ws[f'B{row}']
        if value != '':
            cell.value = value
        cell.style = style
        row += 1
    
    row += 1
    
    # Terminal Value section
    ws[f'A{row}'] = 'Terminal Value Assumptions'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    tv_items = [
        ('Terminal Growth Rate (g)', growth/100, 'percent_style'),
        ('Exit EBITDA Multiple', 8.0, 'input_style'),
        ('Method', 'Perpetuity Growth', 'input_style'),
    ]
    
    for item, value, style in tv_items:
        ws[f'A{row}'] = f'  {item}'
        cell = ws[f'B{row}']
        cell.value = value
        cell.style = style
        row += 1
    
    row += 1
    
    # Other assumptions
    ws[f'A{row}'] = 'Other Assumptions'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    other_items = [
        ('Forecast Period (years)', years, 'input_style'),
        ('Net Debt', 50.0, 'currency_input_style'),
        ('Minority Interest', 0.0, 'currency_input_style'),
        ('Fully Diluted Shares', 10.0, 'input_style'),
    ]
    
    for item, value, style in other_items:
        ws[f'A{row}'] = f'  {item}'
        cell = ws[f'B{row}']
        cell.value = value
        cell.style = style
        row += 1
    
    return row + 2


def add_projections_section(ws, start_row: int, years: int):
    """Add the financial projections section."""
    row = start_row
    
    # Section header
    ws[f'A{row}'] = 'FINANCIAL PROJECTIONS ($ millions)'
    ws[f'A{row}'].style = 'subheader_style'
    ws.merge_cells(f'A{row}:M{row}')
    row += 1
    
    # Year headers
    ws[f'A{row}'] = 'Income Statement'
    ws[f'A{row}'].font = Font(bold=True)
    
    # Historical/Projected labels
    ws[f'B{row}'] = 'LTM'
    ws[f'B{row}'].font = Font(bold=True, italic=True)
    ws[f'B{row}'].alignment = Alignment(horizontal='center')
    
    for i in range(years):
        col = get_column_letter(3 + i)
        ws[f'{col}{row}'] = f'Year {i+1}'
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    ws[f'{get_column_letter(3 + years)}{row}'] = 'Terminal'
    ws[f'{get_column_letter(3 + years)}{row}'].font = Font(bold=True, italic=True)
    ws[f'{get_column_letter(3 + years)}{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    # Income statement line items
    income_items = [
        ('Revenue', 'currency_input_style', 100.0),
        ('Revenue Growth', 'percent_style', 0.10),
        ('', '', ''),
        ('COGS', 'currency_input_style', 60.0),
        ('Gross Profit', 'currency_style', ''),
        ('Gross Margin', 'percent_style', ''),
        ('', '', ''),
        ('Operating Expenses', 'currency_input_style', 20.0),
        ('EBITDA', 'currency_style', ''),
        ('EBITDA Margin', 'percent_style', ''),
        ('', '', ''),
        ('D&A', 'currency_input_style', 5.0),
        ('EBIT', 'currency_style', ''),
        ('EBIT Margin', 'percent_style', ''),
        ('', '', ''),
        ('Interest Expense', 'currency_input_style', 2.0),
        ('EBT', 'currency_style', ''),
        ('Taxes', 'currency_style', ''),
        ('Tax Rate', 'percent_style', ''),
        ('Net Income', 'currency_style', ''),
    ]
    
    for item, style, value in income_items:
        if item == '':
            row += 1
            continue
        
        ws[f'A{row}'] = f'  {item}'
        
        # LTM value
        cell = ws[f'B{row}']
        if value != '':
            cell.value = value
        cell.style = style
        
        # Projected years
        for i in range(years + 1):
            col = get_column_letter(3 + i)
            cell = ws[f'{col}{row}']
            cell.value = ''
            if 'input' in style:
                cell.style = 'input_style' if 'percent' not in style else 'percent_style'
            else:
                cell.style = style
        
        row += 1
    
    return row + 1


def add_free_cash_flow_section(ws, start_row: int, years: int):
    """Add the free cash flow calculation section."""
    row = start_row
    
    # Section header
    ws[f'A{row}'] = 'UNLEVERED FREE CASH FLOW CALCULATION ($ millions)'
    ws[f'A{row}'].style = 'subheader_style'
    ws.merge_cells(f'A{row}:M{row}')
    row += 1
    
    # Year headers
    ws[f'A{row}'] = 'Free Cash Flow Build'
    ws[f'A{row}'].font = Font(bold=True)
    
    ws[f'B{row}'] = 'LTM'
    ws[f'B{row}'].font = Font(bold=True, italic=True)
    ws[f'B{row}'].alignment = Alignment(horizontal='center')
    
    for i in range(years):
        col = get_column_letter(3 + i)
        ws[f'{col}{row}'] = f'Year {i+1}'
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    ws[f'{get_column_letter(3 + years)}{row}'] = 'Terminal'
    ws[f'{get_column_letter(3 + years)}{row}'].font = Font(bold=True, italic=True)
    ws[f'{get_column_letter(3 + years)}{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    # FCF build items
    fcf_items = [
        ('EBIT', 'currency_style'),
        ('(-) Taxes on EBIT', 'currency_style'),
        ('= NOPAT', 'currency_style'),
        ('', ''),
        ('(+)', ''),
        ('D&A', 'currency_style'),
        ('', ''),
        ('(-)', ''),
        ('Capex', 'currency_input_style'),
        ('', ''),
        ('(-)', ''),
        ('Change in NWC', 'currency_input_style'),
        ('NWC % of Revenue', 'percent_input'),
        ('', ''),
        ('= Unlevered Free Cash Flow', 'currency_style'),
    ]
    
    for item, style in fcf_items:
        if item == '':
            row += 1
            continue
        
        ws[f'A{row}'] = f'  {item}'
        if item.startswith('='):
            ws[f'A{row}'].font = Font(bold=True)
        
        for i in range(years + 2):  # LTM + years + Terminal
            col = get_column_letter(2 + i)
            cell = ws[f'{col}{row}']
            cell.value = ''
            if style == 'currency_input_style':
                cell.style = 'currency_input_style'
            elif style == 'percent_input':
                cell.style = 'percent_style'
            elif style:
                cell.style = style
        
        row += 1
    
    return row + 1


def add_valuation_section(ws, start_row: int, years: int):
    """Add the DCF valuation section."""
    row = start_row
    
    # Section header
    ws[f'A{row}'] = 'DCF VALUATION ($ millions)'
    ws[f'A{row}'].style = 'subheader_style'
    ws.merge_cells(f'A{row}:M{row}')
    row += 1
    
    # Year headers
    ws[f'A{row}'] = 'Present Value Calculation'
    ws[f'A{row}'].font = Font(bold=True)
    
    for i in range(years):
        col = get_column_letter(3 + i)
        ws[f'{col}{row}'] = f'Year {i+1}'
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    
    # PV calculation items
    pv_items = [
        ('Unlevered FCF', 'currency_style'),
        ('Discount Factor', 'calc_style'),
        ('Present Value', 'currency_style'),
    ]
    
    for item, style in pv_items:
        ws[f'A{row}'] = f'  {item}'
        if item == 'Present Value':
            ws[f'A{row}'].font = Font(bold=True)
        
        for i in range(years):
            col = get_column_letter(3 + i)
            cell = ws[f'{col}{row}']
            cell.value = ''
            cell.style = style
        
        row += 1
    
    row += 1
    
    # Terminal Value
    ws[f'A{row}'] = 'Terminal Value Calculation'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    tv_items = [
        ('Terminal Year EBITDA', 'currency_style'),
        ('Exit Multiple', 'input_style'),
        ('Terminal Value', 'currency_style'),
        ('', ''),
        ('PV of Terminal Value', 'currency_style'),
    ]
    
    for item, style in tv_items:
        if item == '':
            row += 1
            continue
        
        ws[f'A{row}'] = f'  {item}'
        if item.startswith('PV'):
            ws[f'A{row}'].font = Font(bold=True)
        
        cell = ws[f'B{row}']
        cell.value = ''
        cell.style = style
        
        row += 1
    
    row += 1
    
    # Enterprise to Equity bridge
    bridge_items = [
        ('ENTERPRISE VALUE', 'highlight_style'),
        ('', ''),
        ('PV of Explicit Period FCF', 'currency_style'),
        ('(+) PV of Terminal Value', 'currency_style'),
        ('= Enterprise Value', 'highlight_style'),
        ('', ''),
        ('(+)', ''),
        ('Cash & Equivalents', 'currency_input_style'),
        ('(-)', ''),
        ('Total Debt', 'currency_input_style'),
        ('(-)', ''),
        ('Minority Interest', 'currency_input_style'),
        ('= EQUITY VALUE', 'highlight_style'),
        ('', ''),
        ('Fully Diluted Shares Outstanding', 'input_style'),
        ('Implied Share Price', 'highlight_style'),
    ]
    
    for item, style in bridge_items:
        if item == '':
            row += 1
            continue
        
        ws[f'A{row}'] = item
        if item.isupper() or 'Value' in item:
            ws[f'A{row}'].font = Font(bold=True)
        elif item in ['(+)', '(-)']:
            ws[f'A{row}'].font = Font(bold=True)
        else:
            ws[f'A{row}'] = f'  {item}'
        
        cell = ws[f'B{row}']
        cell.value = ''
        cell.style = style
        
        row += 1
    
    return row + 1


def add_sensitivity_section(ws, start_row: int):
    """Add sensitivity analysis section."""
    row = start_row
    
    # Section header
    ws[f'A{row}'] = 'SENSITIVITY ANALYSIS'
    ws[f'A{row}'].style = 'subheader_style'
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # WACC vs Terminal Growth sensitivity
    ws[f'A{row}'] = 'Implied Enterprise Value Sensitivity ($ millions)'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    ws[f'A{row}'] = 'WACC / Terminal Growth'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['header_fill'],
                                      end_color=COLORS['header_fill'], fill_type='solid')
    
    # Growth rate headers
    growth_rates = ['1.0%', '1.5%', '2.0%', '2.5%', '3.0%']
    for col, rate in enumerate(growth_rates, 2):
        cell = ws.cell(row=row, column=col, value=rate)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['header_fill'],
                                end_color=COLORS['header_fill'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # WACC rows
    waccs = ['7.5%', '8.5%', '9.5%', '10.5%', '11.5%']
    for wacc in waccs:
        ws[f'A{row}'] = wacc
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color=COLORS['header_fill'],
                                         end_color=COLORS['header_fill'], fill_type='solid')
        
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col, value='')
            cell.style = 'currency_style'
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
    
    return row + 2


def add_notes_section(ws, start_row: int):
    """Add notes and methodology section."""
    row = start_row
    
    ws[f'A{row}'] = 'METHODOLOGY & NOTES'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 2
    
    notes = [
        "Free Cash Flow Formula:",
        "  UFCF = EBIT(1 - Tax Rate) + D&A - Capex - ΔNWC",
        "",
        "Terminal Value (Perpetuity Growth):",
        "  TV = UFCF(n+1) / (WACC - g) = UFCF(n) × (1 + g) / (WACC - g)",
        "",
        "Present Value Formula:",
        "  PV = FCF / (1 + WACC)^n",
        "",
        "WACC Formula:",
        "  WACC = (E/V) × Re + (D/V) × Rd × (1 - Tc)",
        "  Re = Rf + β × (Market Risk Premium)",
        "",
        "Color Coding:",
        "  - Blue text = INPUTS (hardcoded assumptions)",
        "  - Black text = CALCULATIONS (formulas)",
        "",
        "Important Notes:",
        "  - All values in $ millions unless otherwise noted",
        "  - Use fully diluted share count for equity value",
        "  - Calendarize fiscal years if necessary",
        "  - Sensitivity analysis shows range of outcomes",
        "  - Always footnote key assumptions",
    ]
    
    for note in notes:
        ws[f'A{row}'] = note
        if note.endswith(':') and not note.startswith('  '):
            ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    return row


def main():
    parser = argparse.ArgumentParser(
        description='Generate DCF Valuation Excel Model',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    %(prog)s --company "Acme Corp" --years 5 --output acme_dcf.xlsx
    %(prog)s --company "Acme Corp" --wacc 9.5 --growth 2.5 --years 10 --output dcf.xlsx
        """
    )
    
    parser.add_argument('--company', '-c', required=True,
                        help='Company name')
    parser.add_argument('--years', '-y', type=int, default=5,
                        help='Number of forecast years (default: 5)')
    parser.add_argument('--wacc', '-w', type=float, default=9.5,
                        help='WACC in percent (default: 9.5)')
    parser.add_argument('--growth', '-g', type=float, default=2.5,
                        help='Terminal growth rate in percent (default: 2.5)')
    parser.add_argument('--output', '-o', default='dcf_valuation.xlsx',
                        help='Output Excel file name (default: dcf_valuation.xlsx)')
    
    args = parser.parse_args()
    
    # Validate inputs
    if args.years < 3 or args.years > 10:
        print("Error: Years must be between 3 and 10")
        sys.exit(1)
    
    if args.wacc < 1 or args.wacc > 30:
        print("Error: WACC should be between 1% and 30%")
        sys.exit(1)
    
    if args.growth < 0 or args.growth > 10:
        print("Error: Growth rate should be between 0% and 10%")
        sys.exit(1)
    
    print(f"Generating DCF Model for: {args.company}")
    print(f"Forecast Years: {args.years}")
    print(f"WACC: {args.wacc}%")
    print(f"Terminal Growth: {args.growth}%")
    print(f"Output: {args.output}")
    
    # Create workbook
    wb = Workbook()
    create_styles(wb)
    
    # Setup main sheet
    ws = wb.active
    setup_sheet(ws, 'DCF Model')
    
    # Add sections
    current_row = 3
    current_row = add_assumptions_section(ws, current_row, args.wacc, args.growth, args.years)
    current_row = add_projections_section(ws, current_row, args.years)
    current_row = add_free_cash_flow_section(ws, current_row, args.years)
    current_row = add_valuation_section(ws, current_row, args.years)
    current_row = add_sensitivity_section(ws, current_row)
    add_notes_section(ws, current_row)
    
    # Save
    wb.save(args.output)
    print(f"✓ Generated: {args.output}")
    print("\nNext steps:")
    print("1. Open the Excel file")
    print("2. Fill in historical financial data (LTM column)")
    print("3. Adjust assumptions (blue cells)")
    print("4. Build projection formulas")
    print("5. Review sensitivity analysis")


if __name__ == '__main__':
    main()
