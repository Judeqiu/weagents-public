#!/usr/bin/env python3
"""
Generate Comparable Company Analysis (Comps) Excel Model

Usage:
    python3 generate-comps.py --target "Company Name" --peers "Peer1,Peer2,Peer3" --output comps.xlsx
    python3 generate-comps.py --target "Company Name" --peers-file peers.csv --output comps.xlsx
"""

import argparse
import csv
import sys
from pathlib import Path
from typing import List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)


# Financial color scheme (investment banking standard)
COLORS = {
    'input_blue': '0000FF',      # Blue - hardcoded inputs
    'calc_black': '000000',      # Black - formulas/calculations
    'link_green': '008000',      # Green - links to other sheets
    'header_fill': 'D9E1F2',     # Light blue for headers
    'subheader_fill': 'E7E6E6',  # Light gray for subheaders
    'alternate_fill': 'F2F2F2',  # Very light gray for alternating rows
    'border': '000000',          # Black borders
}


def create_styles(wb: Workbook):
    """Create named styles for the workbook."""
    
    # Input style (blue text - hardcoded values)
    input_style = NamedStyle(name='input_style')
    input_style.font = Font(color=COLORS['input_blue'], size=10)
    input_style.number_format = '#,##0.0'
    wb.add_named_style(input_style)
    
    # Calc style (black text - formulas)
    calc_style = NamedStyle(name='calc_style')
    calc_style.font = Font(color=COLORS['calc_black'], size=10)
    calc_style.number_format = '#,##0.0'
    wb.add_named_style(calc_style)
    
    # Multiple style (x suffix)
    multiple_style = NamedStyle(name='multiple_style')
    multiple_style.font = Font(color=COLORS['calc_black'], size=10)
    multiple_style.number_format = '0.0"x"'
    wb.add_named_style(multiple_style)
    
    # Percent style
    percent_style = NamedStyle(name='percent_style')
    percent_style.font = Font(color=COLORS['calc_black'], size=10)
    percent_style.number_format = '0.0%'
    wb.add_named_style(percent_style)
    
    # Currency style (millions)
    currency_style = NamedStyle(name='currency_style')
    currency_style.font = Font(color=COLORS['calc_black'], size=10)
    currency_style.number_format = '$#,##0.0,,"M"'
    wb.add_named_style(currency_style)
    
    # Header style
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(bold=True, color='FFFFFF', size=11)
    header_style.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wb.add_named_style(header_style)
    
    # Subheader style
    subheader_style = NamedStyle(name='subheader_style')
    subheader_style.font = Font(bold=True, size=10)
    subheader_style.fill = PatternFill(start_color=COLORS['subheader_fill'], 
                                       end_color=COLORS['subheader_fill'], fill_type='solid')
    wb.add_named_style(subheader_style)
    
    return input_style, calc_style


def setup_sheet(ws, title: str):
    """Setup worksheet with title and basic formatting."""
    ws.title = title
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Title
    ws['A1'] = 'COMPARABLE COMPANY ANALYSIS'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')


def add_company_data_section(ws, start_row: int, target_name: str, peers: List[str]):
    """Add the company data input section."""
    row = start_row
    
    # Section header
    ws[f'A{row}'] = 'COMPANY DATA INPUTS'
    ws[f'A{row}'].style = 'subheader_style'
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Column headers
    headers = ['Company', 'Share Price', 'Shares Outstanding', 'Market Cap', 
               'Total Debt', 'Cash', 'Enterprise Value', 'LTM EBITDA']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.style = 'header_style'
    row += 1
    
    # Target company row
    ws[f'A{row}'] = f'{target_name} (Target)'
    ws[f'A{row}'].font = Font(bold=True)
    
    # Add placeholder values with blue text (input style)
    for col in range(2, 9):
        cell = ws.cell(row=row, column=col, value=0.0)
        cell.style = 'input_style'
    row += 1
    
    # Peer rows
    for peer in peers:
        ws[f'A{row}'] = peer
        for col in range(2, 9):
            cell = ws.cell(row=row, column=col, value=0.0)
            cell.style = 'input_style'
        row += 1
    
    # Add statistical rows
    row += 1
    ws[f'A{row}'] = 'MEDIAN'
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    row += 1
    
    ws[f'A{row}'] = 'MEAN'
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    row += 1
    
    ws[f'A{row}'] = 'HIGH'
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    row += 1
    
    ws[f'A{row}'] = 'LOW'
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    return row + 2


def add_valuation_multiples_section(ws, start_row: int, peers: List[str]):
    """Add the valuation multiples section."""
    row = start_row
    
    # Section header
    ws[f'A{row}'] = 'VALUATION MULTIPLES'
    ws[f'A{row}'].style = 'subheader_style'
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Enterprise Value Multiples
    ws[f'A{row}'] = 'Enterprise Value Multiples'
    ws[f'A{row}'].font = Font(bold=True, size=10)
    row += 1
    
    headers = ['Company', 'EV/Revenue', 'EV/EBITDA', 'EV/EBIT']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.style = 'header_style'
    row += 1
    
    # Data rows (placeholder formulas)
    num_peers = len(peers) + 1  # +1 for target
    for i in range(num_peers):
        ws.cell(row=row, column=1, value='' if i > 0 else 'Target')  # Company name
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col, value='')
            cell.style = 'multiple_style'
        row += 1
    
    # Statistics
    row += 1
    for stat in ['MEDIAN', 'MEAN', 'HIGH', 'LOW']:
        ws[f'A{row}'] = stat
        ws[f'A{row}'].font = Font(bold=True, italic=True)
        row += 1
    
    row += 1
    
    # Equity Value Multiples
    ws[f'A{row}'] = 'Equity Value Multiples'
    ws[f'A{row}'].font = Font(bold=True, size=10)
    row += 1
    
    headers = ['Company', 'P/E', 'P/B', 'Dividend Yield']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.style = 'header_style'
    row += 1
    
    # Data rows
    for i in range(num_peers):
        ws.cell(row=row, column=1, value='')
        for col in range(2, 4):
            cell = ws.cell(row=row, column=col, value='')
            cell.style = 'multiple_style'
        # Dividend yield as percent
        cell = ws.cell(row=row, column=4, value='')
        cell.style = 'percent_style'
        row += 1
    
    # Statistics
    row += 1
    for stat in ['MEDIAN', 'MEAN', 'HIGH', 'LOW']:
        ws[f'A{row}'] = stat
        ws[f'A{row}'].font = Font(bold=True, italic=True)
        row += 1
    
    return row + 1


def add_implied_valuation_section(ws, start_row: int):
    """Add the implied valuation section for target company."""
    row = start_row
    
    # Section header
    ws[f'A{row}'] = 'IMPLIED VALUATION (Target Company)'
    ws[f'A{row}'].style = 'subheader_style'
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Column headers
    headers = ['Metric', 'Low', 'Median', 'High', 'Implied EV']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.style = 'header_style'
    row += 1
    
    # EV/EBITDA valuation
    ws[f'A{row}'] = 'EV/EBITDA'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(2, 6):
        cell = ws.cell(row=row, column=col, value='')
        if col == 5:
            cell.style = 'currency_style'
        else:
            cell.style = 'multiple_style'
    row += 1
    
    # EV/Revenue valuation
    ws[f'A{row}'] = 'EV/Revenue'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(2, 6):
        cell = ws.cell(row=row, column=col, value='')
        if col == 5:
            cell.style = 'currency_style'
        else:
            cell.style = 'multiple_style'
    row += 1
    
    # P/E valuation
    ws[f'A{row}'] = 'P/E (Equity Value)'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(2, 6):
        cell = ws.cell(row=row, column=col, value='')
        if col == 5:
            cell.style = 'currency_style'
        else:
            cell.style = 'multiple_style'
    row += 2
    
    # Summary
    ws[f'A{row}'] = 'VALUATION SUMMARY'
    ws[f'A{row}'].style = 'subheader_style'
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    summary_items = [
        ('Implied Enterprise Value Range', '', 'currency_style'),
        ('(+) Cash', '', 'currency_style'),
        ('(-) Total Debt', '', 'currency_style'),
        ('= Implied Equity Value Range', '', 'currency_style'),
        ('', '', ''),
        ('Implied Share Price Range', '', 'currency_style'),
    ]
    
    for item, _, style_name in summary_items:
        if item:
            ws[f'A{row}'] = item
            ws[f'A{row}'].font = Font(bold=True)
            cell = ws[f'E{row}']
            cell.value = ''
            cell.style = style_name
        row += 1
    
    return row


def add_notes_section(ws, start_row: int):
    """Add notes and methodology section."""
    row = start_row
    
    ws[f'A{row}'] = 'NOTES & METHODOLOGY'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 2
    
    notes = [
        "Key Definitions:",
        "- Enterprise Value (EV) = Market Cap + Total Debt - Cash",
        "- LTM = Last Twelve Months",
        "- NTM = Next Twelve Months (forward-looking)",
        "",
        "Formula Reference:",
        "- All blue text cells are INPUTS (hardcoded values)",
        "- All black text cells are CALCULATIONS (formulas)",
        "- EV multiples use Enterprise Value in numerator",
        "- Equity multiples use Market Cap in numerator",
        "",
        "Calculation Notes:",
        "- Share counts should be fully diluted",
        "- Calendarize fiscal years if different from peers",
        "- Adjust for non-recurring items in EBITDA",
        "- Use most recent balance sheet for debt/cash",
    ]
    
    for note in notes:
        ws[f'A{row}'] = note
        if note.endswith(':'):
            ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    return row


def parse_peers(peers_str: str) -> List[str]:
    """Parse comma-separated peer list."""
    return [p.strip() for p in peers_str.split(',') if p.strip()]


def load_peers_from_file(filepath: str) -> List[str]:
    """Load peers from CSV file."""
    peers = []
    with open(filepath, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            if row and row[0].strip():
                peers.append(row[0].strip())
    return peers


def main():
    parser = argparse.ArgumentParser(
        description='Generate Comparable Company Analysis Excel Model',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    %(prog)s --target "Acme Corp" --peers "PeerA,PeerB,PeerC" --output comps.xlsx
    %(prog)s --target "Acme Corp" --peers-file peers.csv --output comps.xlsx
        """
    )
    
    parser.add_argument('--target', '-t', required=True,
                        help='Target company name')
    parser.add_argument('--peers', '-p',
                        help='Comma-separated list of peer company names')
    parser.add_argument('--peers-file', '-f',
                        help='CSV file containing peer names (one per row)')
    parser.add_argument('--output', '-o', default='comps_analysis.xlsx',
                        help='Output Excel file name (default: comps_analysis.xlsx)')
    
    args = parser.parse_args()
    
    # Validate inputs
    if not args.peers and not args.peers_file:
        print("Error: Must provide either --peers or --peers-file")
        sys.exit(1)
    
    # Load peers
    if args.peers_file:
        peers = load_peers_from_file(args.peers_file)
    else:
        peers = parse_peers(args.peers)
    
    if not peers:
        print("Error: No peers specified")
        sys.exit(1)
    
    print(f"Generating Comps Analysis for: {args.target}")
    print(f"Peers ({len(peers)}): {', '.join(peers)}")
    print(f"Output: {args.output}")
    
    # Create workbook
    wb = Workbook()
    create_styles(wb)
    
    # Setup main sheet
    ws = wb.active
    setup_sheet(ws, 'Comps Analysis')
    
    # Add sections
    current_row = 3
    current_row = add_company_data_section(ws, current_row, args.target, peers)
    current_row = add_valuation_multiples_section(ws, current_row, peers)
    current_row = add_implied_valuation_section(ws, current_row)
    add_notes_section(ws, current_row)
    
    # Add gridlines
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )
    
    # Save
    wb.save(args.output)
    print(f"✓ Generated: {args.output}")
    print("\nNext steps:")
    print("1. Open the Excel file")
    print("2. Fill in company data (blue cells)")
    print("3. Add formulas in calculation cells")
    print("4. Review implied valuation ranges")


if __name__ == '__main__':
    main()
